"""Script to upload Excel data to Google Sheets with Week 1 and Week 2 support."""
import sys
import os

# Add the backend directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from app.services.google.sheets_client import GoogleSheetsClient

# Day names that indicate actual delivery data (not headers/totals)
VALID_DAYS = {"monday", "tuesday", "wednesday", "thursday", "friday"}


def read_excel_sheet(workbook, sheet_name: str) -> list:
    """Read data from an Excel sheet and return as 2D list."""
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in workbook")
        return []

    sheet = workbook[sheet_name]
    data = []

    for row in sheet.iter_rows(values_only=True):
        # Convert None values to empty strings for Google Sheets
        row_data = [cell if cell is not None else "" for cell in row]
        # Only include rows that have some data in first 7 columns
        if any(cell != "" for cell in row_data[:7]):
            data.append(row_data[:7])  # Only take columns A-G

    return data


def add_week2_data(data: list) -> list:
    """Duplicate Week 1 delivery rows as Week 2 with 'W2' suffix on day names.

    Args:
        data: Original data with headers and Week 1 deliveries

    Returns:
        Data with Week 2 rows appended
    """
    week2_rows = []

    for row in data:
        if not row or not row[0]:
            continue

        day_name = str(row[0]).lower().strip()

        # Check if this is a delivery row (has a valid day name)
        if day_name in VALID_DAYS:
            # Check if it has SQM data (column F, index 5)
            try:
                sqm = float(row[5]) if row[5] else 0
            except (ValueError, TypeError):
                sqm = 0

            if sqm > 0:
                # Create Week 2 version with "W2" suffix
                week2_row = list(row)
                week2_row[0] = f"{row[0]} W2"
                week2_rows.append(week2_row)

    # Append Week 2 rows to the data
    if week2_rows:
        # Add a blank row separator
        data.append(["", "", "", "", "", "", ""])
        # Add Week 2 header
        data.append(["WEEK 2", "", "", "", "", "", ""])
        # Add Week 2 data
        data.extend(week2_rows)

    return data


def main():
    excel_path = r"c:\Users\waqar\OneDrive\Desktop\Projects\Learning\Rob\Turf Supply Tracker 2 Trucks.xlsx"

    print(f"Loading Excel file: {excel_path}")
    workbook = load_workbook(excel_path, data_only=True)

    print(f"Available sheets: {workbook.sheetnames}")

    # Read data from both truck sheets
    truck1_data = read_excel_sheet(workbook, "Truck 1")
    truck2_data = read_excel_sheet(workbook, "Truck 2")

    print(f"\nTruck 1 data: {len(truck1_data)} rows (Week 1)")
    print(f"Truck 2 data: {len(truck2_data)} rows (Week 1)")

    # Add Week 2 data (duplicate of Week 1)
    print("\nAdding Week 2 data (duplicate of Week 1)...")
    truck1_data = add_week2_data(truck1_data)
    truck2_data = add_week2_data(truck2_data)

    print(f"Truck 1 data: {len(truck1_data)} rows (with Week 2)")
    print(f"Truck 2 data: {len(truck2_data)} rows (with Week 2)")

    # Upload to Google Sheets
    print("\nConnecting to Google Sheets...")
    client = GoogleSheetsClient()

    # Check available sheets in Google
    available = client.get_available_sheets()
    print(f"Google Sheets tabs: {available}")

    # Upload Truck 1 data
    if truck1_data and "Truck 1" in available:
        print("\nUploading Truck 1 data (Week 1 + Week 2)...")
        cells_updated = client.update_worksheet_data("Truck 1", truck1_data)
        print(f"  Updated {cells_updated} cells")

    # Upload Truck 2 data
    if truck2_data and "Truck 2" in available:
        print("\nUploading Truck 2 data (Week 1 + Week 2)...")
        cells_updated = client.update_worksheet_data("Truck 2", truck2_data)
        print(f"  Updated {cells_updated} cells")

    print("\nDone! The Google Sheet has been updated with Week 1 and Week 2 data.")
    print("Refresh the dashboard to see the new data.")


if __name__ == "__main__":
    main()
